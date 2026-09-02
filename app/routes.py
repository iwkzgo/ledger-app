import io
from datetime import datetime
from urllib.parse import quote

from flask import Blueprint, Response, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference

from .aggregator import (
    build_budget_status,
    build_calendar_month,
    build_expense_forecast,
    build_monthly_summary,
    build_savings_goal_status,
    build_weekly_pattern,
    current_month,
)
from .categories import ensure_default_categories, get_expense_categories
from .categorizer import categorize_item, learn_category
from .chart import build_line_chart
from .extensions import db
from .models import Budget, Category, Entry, LearnedCategory, SavingsGoal
from .recurring import sync_recurring_items
from .timeutil import to_kst
from .rules import (
    FALLBACK_CATEGORY,
    FIXED_EXPENSE_CATEGORY,
    INCOME_CATEGORY,
    SAVINGS_CATEGORY,
)
from .text_parser import parse_amount, parse_entry

bp = Blueprint("ledger", __name__)


def category_choices_for(user_id: int):
    return [INCOME_CATEGORY, SAVINGS_CATEGORY] + get_expense_categories(user_id)


def flash_budget_warning(category: str) -> None:
    if category in (INCOME_CATEGORY, SAVINGS_CATEGORY):
        return
    for status in build_budget_status(current_user.id):
        if status["category"] != category:
            continue
        if status["level"] == "over":
            flash(
                f'⚠️ "{category}" 예산을 초과했습니다! '
                f'({status["spent"]:,}원 / {status["budget"]:,}원, {status["percent"]}%)',
                "danger",
            )
        elif status["level"] == "warning":
            flash(
                f'"{category}" 예산의 {status["percent"]}%를 사용했습니다. '
                f'({status["spent"]:,}원 / {status["budget"]:,}원)',
                "warning",
            )
        break


@bp.route("/")
@login_required
def index():
    recorded = sync_recurring_items(current_user.id)
    if recorded:
        flash("고정 지출/수입이 자동 기록되었습니다: " + ", ".join(recorded), "info")
        flash_budget_warning(FIXED_EXPENSE_CATEGORY)

    recent_entries = (
        Entry.query.filter_by(user_id=current_user.id)
        .order_by(Entry.id.desc())
        .limit(10)
        .all()
    )
    weekly_pattern = build_weekly_pattern(current_user.id)
    return render_template(
        "index.html",
        summary=build_monthly_summary(current_user.id),
        recent_entries=recent_entries,
        budget_status=build_budget_status(current_user.id),
        current_month=current_month(),
        savings_goal=build_savings_goal_status(current_user.id, current_month()),
        weekly_chart=build_line_chart(weekly_pattern),
        weekly_total=sum(day["amount"] for day in weekly_pattern),
        expense_forecast=build_expense_forecast(current_user.id),
    )


@bp.route("/entry", methods=["POST"])
@login_required
def create_entry():
    raw_text = request.form.get("raw_text", "").strip()
    if not raw_text:
        flash("입력할 내용이 없습니다.", "error")
        return redirect(url_for("ledger.index"))

    parsed = parse_entry(raw_text)
    if parsed is None:
        flash(f'"{raw_text}"에서 금액을 찾지 못했어요. 예: 스벅 5천원', "error")
        return redirect(url_for("ledger.index"))

    item, amount = parsed
    category = categorize_item(item, current_user.id)
    valid_categories = set(get_expense_categories(current_user.id)) | {INCOME_CATEGORY, SAVINGS_CATEGORY}
    if category not in valid_categories:
        # 규칙이 가리키는 카테고리를 사용자가 삭제한 경우, 미분류로 남기지 않고 안전하게 기타로 보냅니다.
        category = FALLBACK_CATEGORY

    if category == FALLBACK_CATEGORY:
        return render_template(
            "confirm_category.html",
            raw_text=raw_text,
            item=item,
            amount=amount,
            categories=category_choices_for(current_user.id),
        )

    db.session.add(
        Entry(user_id=current_user.id, raw_text=raw_text, item=item, amount=amount, category=category)
    )
    db.session.commit()
    learn_category(item, category, current_user.id)
    flash(f'저장했습니다: {item} · {amount:,}원 · {category}', "success")
    flash_budget_warning(category)
    return redirect(url_for("ledger.index"))


@bp.route("/entry/confirm", methods=["POST"])
@login_required
def confirm_entry():
    raw_text = request.form.get("raw_text", "")
    item = request.form.get("item", "")
    amount = int(request.form.get("amount", "0"))
    category = request.form.get("category", FALLBACK_CATEGORY)

    db.session.add(
        Entry(user_id=current_user.id, raw_text=raw_text, item=item, amount=amount, category=category)
    )
    db.session.commit()
    learn_category(item, category, current_user.id)
    flash(f'저장했습니다: {item} · {amount:,}원 · {category}', "success")
    flash_budget_warning(category)
    return redirect(url_for("ledger.index"))


@bp.route("/entry/<int:entry_id>/delete", methods=["POST"])
@login_required
def delete_entry(entry_id):
    entry = Entry.query.filter_by(id=entry_id, user_id=current_user.id).first_or_404()
    db.session.delete(entry)
    db.session.commit()
    flash("삭제했습니다.", "info")
    return redirect(url_for("ledger.index"))


@bp.route("/entry/<int:entry_id>/edit", methods=["GET", "POST"])
@login_required
def edit_entry(entry_id):
    entry = Entry.query.filter_by(id=entry_id, user_id=current_user.id).first_or_404()
    next_url = request.values.get("next") or url_for("ledger.index")
    if not next_url.startswith("/"):
        next_url = url_for("ledger.index")

    choices = category_choices_for(current_user.id)

    if request.method == "POST":
        item = request.form.get("item", "").strip()
        amount_raw = request.form.get("amount", "").strip()
        category = request.form.get("category", "").strip()
        try:
            amount = int(amount_raw)
        except ValueError:
            amount = None

        if not item or not amount or amount <= 0 or category not in choices:
            flash("항목, 금액, 카테고리를 올바르게 입력해주세요.", "error")
            return render_template(
                "edit_entry.html",
                entry=entry,
                next_url=next_url,
                form_item=item,
                form_amount=amount_raw,
                form_category=category or entry.category,
                category_choices=choices,
            )

        entry.item = item
        entry.amount = amount
        entry.category = category
        db.session.commit()
        learn_category(item, category, current_user.id)
        flash(f'수정했습니다: {item} · {amount:,}원 · {category}', "success")
        return redirect(next_url)

    return render_template(
        "edit_entry.html",
        entry=entry,
        next_url=next_url,
        form_item=entry.item,
        form_amount=entry.amount,
        form_category=entry.category,
        category_choices=choices,
    )


@bp.route("/savings-goal", methods=["POST"])
@login_required
def savings_goal():
    month = request.form.get("month", "").strip() or current_month()
    amount = parse_amount(request.form.get("target_amount", "").strip()) or 0

    goal = SavingsGoal.query.filter_by(user_id=current_user.id, month=month).first()
    if goal is None:
        db.session.add(SavingsGoal(user_id=current_user.id, month=month, target_amount=amount))
    else:
        goal.target_amount = amount
    db.session.commit()
    flash(f'{month} 저축 목표를 {amount:,}원으로 설정했습니다.', "success")
    return redirect(url_for("ledger.index"))


@bp.route("/export")
@login_required
def export_csv():
    month = request.args.get("month", "").strip() or None

    query = Entry.query.filter_by(user_id=current_user.id)
    filename = "전체_가계부.xlsx"
    if month:
        try:
            year, mon = map(int, month.split("-"))
            start = datetime(year, mon, 1)
            end = datetime(year + 1, 1, 1) if mon == 12 else datetime(year, mon + 1, 1)
            query = query.filter(Entry.created_at >= start, Entry.created_at < end)
            filename = f"{year}년_{mon}월_가계부.xlsx"
        except ValueError:
            flash(f'"{month}"은(는) 올바른 월 형식이 아니에요.', "error")
            return redirect(url_for("ledger.index"))

    entries = query.order_by(Entry.created_at.asc()).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "가계부"
    sheet.append(["날짜", "항목", "금액", "카테고리"])
    for entry in entries:
        sheet.append(
            [to_kst(entry.created_at).strftime("%Y-%m-%d"), entry.item, entry.amount, entry.category]
        )
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 12
    sheet.column_dimensions["D"].width = 12

    # 지출 카테고리별 합계 (수입/저축 제외) - 원형 그래프의 재료가 되는 표.
    # F, G열에 같은 시트 안에 넣어서 거래내역과 한 화면에서 같이 보이게 합니다.
    category_totals = {}
    for entry in entries:
        if entry.category in (INCOME_CATEGORY, SAVINGS_CATEGORY):
            continue
        category_totals[entry.category] = category_totals.get(entry.category, 0) + entry.amount

    if category_totals:
        sheet["F1"] = "카테고리"
        sheet["G1"] = "합계"
        sheet.column_dimensions["F"].width = 12
        sheet.column_dimensions["G"].width = 12
        for offset, (category, total) in enumerate(category_totals.items()):
            row = offset + 2
            sheet.cell(row=row, column=6, value=category)
            sheet.cell(row=row, column=7, value=total)

        chart = PieChart()
        chart.title = "카테고리별 지출 비율"
        chart.title.overlay = False  # 제목이 도넛과 겹치지 않도록 별도 공간 확보
        row_count = len(category_totals)
        data_ref = Reference(sheet, min_col=7, min_row=1, max_row=row_count + 1)
        labels_ref = Reference(sheet, min_col=6, min_row=2, max_row=row_count + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        chart.width = 14
        chart.height = 10
        sheet.add_chart(chart, "I2")

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    # 한글 파일명이 깨지지 않도록 RFC 5987 형식(filename*)을 같이 내려줍니다.
    encoded_filename = quote(filename)
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=ledger.xlsx; filename*=UTF-8''{encoded_filename}"
            )
        },
    )


@bp.route("/calendar")
@login_required
def calendar_page():
    month = request.args.get("month", "").strip() or current_month()
    try:
        year, mon = map(int, month.split("-"))
        datetime(year, mon, 1)
    except ValueError:
        month = current_month()

    return render_template("calendar.html", calendar_data=build_calendar_month(current_user.id, month))


@bp.route("/entries")
@login_required
def entries_page():
    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    month = request.args.get("month", "").strip()
    page = request.args.get("page", 1, type=int) or 1

    query = Entry.query.filter_by(user_id=current_user.id)
    if q:
        query = query.filter(Entry.item.ilike(f"%{q}%"))
    if category:
        query = query.filter(Entry.category == category)
    if month:
        try:
            year, mon = map(int, month.split("-"))
            start = datetime(year, mon, 1)
            end = datetime(year + 1, 1, 1) if mon == 12 else datetime(year, mon + 1, 1)
            query = query.filter(Entry.created_at >= start, Entry.created_at < end)
        except ValueError:
            flash(f'"{month}"은(는) 올바른 월 형식이 아니에요.', "error")
            month = ""

    total_amount = query.with_entities(db.func.coalesce(db.func.sum(Entry.amount), 0)).scalar()

    pagination = query.order_by(Entry.created_at.desc()).paginate(
        page=page, per_page=30, error_out=False
    )

    return render_template(
        "entries.html",
        pagination=pagination,
        entries=pagination.items,
        category_choices=category_choices_for(current_user.id),
        q=q,
        selected_category=category,
        month=month,
        total_amount=total_amount,
    )


@bp.route("/budgets", methods=["GET", "POST"])
@login_required
def budgets_page():
    expense_categories = get_expense_categories(current_user.id)
    if request.method == "POST":
        for category in expense_categories:
            raw_value = request.form.get(f"budget__{category}", "").strip()
            amount = parse_amount(raw_value) or 0
            budget = Budget.query.filter_by(user_id=current_user.id, category=category).first()
            if budget is None:
                db.session.add(Budget(user_id=current_user.id, category=category, monthly_amount=amount))
            else:
                budget.monthly_amount = amount
        db.session.commit()
        flash("예산을 저장했습니다.", "success")
        return redirect(url_for("ledger.index"))

    budgets = {
        b.category: b.monthly_amount
        for b in Budget.query.filter_by(user_id=current_user.id).all()
    }
    return render_template("budgets.html", categories=expense_categories, budgets=budgets)


@bp.route("/categories")
@login_required
def categories_page():
    ensure_default_categories(current_user.id)
    categories = Category.query.filter_by(user_id=current_user.id).order_by(Category.position).all()
    return render_template("categories.html", categories=categories)


@bp.route("/categories/add", methods=["POST"])
@login_required
def add_category():
    name = request.form.get("name", "").strip()
    if not name:
        flash("카테고리 이름을 입력해주세요.", "error")
        return redirect(url_for("ledger.categories_page"))
    if name in (INCOME_CATEGORY, SAVINGS_CATEGORY):
        flash(f'"{name}"은(는) 별도로 관리되는 카테고리라 추가할 수 없어요.', "error")
        return redirect(url_for("ledger.categories_page"))
    if Category.query.filter_by(user_id=current_user.id, name=name).first() is not None:
        flash(f'"{name}" 카테고리가 이미 있어요.', "error")
        return redirect(url_for("ledger.categories_page"))

    max_position = (
        db.session.query(db.func.max(Category.position)).filter_by(user_id=current_user.id).scalar() or 0
    )
    db.session.add(Category(user_id=current_user.id, name=name, position=max_position + 1))
    db.session.commit()
    flash(f'"{name}" 카테고리를 추가했습니다.', "success")
    return redirect(url_for("ledger.categories_page"))


@bp.route("/categories/<int:category_id>/delete", methods=["POST"])
@login_required
def delete_category(category_id):
    category = Category.query.filter_by(id=category_id, user_id=current_user.id).first_or_404()
    if category.is_protected:
        flash(f'"{category.name}"은(는) 삭제할 수 없는 카테고리예요.', "error")
        return redirect(url_for("ledger.categories_page"))

    Entry.query.filter_by(user_id=current_user.id, category=category.name).update(
        {"category": FALLBACK_CATEGORY}
    )
    LearnedCategory.query.filter_by(user_id=current_user.id, category=category.name).update(
        {"category": FALLBACK_CATEGORY}
    )
    Budget.query.filter_by(user_id=current_user.id, category=category.name).delete()
    db.session.delete(category)
    db.session.commit()
    flash(f'"{category.name}" 카테고리를 삭제했습니다. 관련 거래내역은 "기타"로 옮겨졌습니다.', "info")
    return redirect(url_for("ledger.categories_page"))
